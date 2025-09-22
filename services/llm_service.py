# services/llm_service.py - LLM integration service
from typing import Optional, Dict, Any
import asyncio
from langchain_google_genai import ChatGoogleGenerativeAI
#from langchain_openai import ChatOpenAI
from langchain_community.llms.anthropic import Anthropic
from langchain_core.prompts import PromptTemplate
from langchain_core.output_parsers import JsonOutputParser
from models.beneish_models import FinancialData
from utils.config import Config
import PyPDF2
import pandas as pd
import openpyxl

class LLMService:
    def __init__(self, config: Config):
        self.config = config
        self.current_provider = None
        self.current_model = None
        self.llm = None
        
    def initialize_llm(self, provider: str, model: str, api_key: str = None) -> bool:
        """Initialize LLM with specified provider and model"""
        try:
            if api_key is None:
                api_key = self.config.get_api_key(provider)
            
            if not api_key:
                raise ValueError(f"No API key found for {provider}")
            
            # if provider == "openai":
            #     # Map to actual available models
            #     api_model_map = {
            #         "gpt-5": "gpt-4-turbo-preview",  # Use latest available
            #         "gpt-4.1": "gpt-4-turbo-preview",
            #         "gpt-4.1-mini": "gpt-4-turbo-preview",
            #         "gpt-4.1-nano": "gpt-3.5-turbo",
            #         "o3": "gpt-4-turbo-preview",
            #         "o4-mini": "gpt-4-turbo-preview"
            #     }
            #     actual_model = api_model_map.get(model, "gpt-4-turbo-preview")
                
            #     self.llm = ChatOpenAI(
            #         model=actual_model,
            #         api_key=api_key,
            #         temperature=0.5
            #     )
            if provider == "anthropic":
                # Map our model names to actual API model names
                api_model_map = {
                    "claude-opus-4.1": "claude-3-opus-20240229",  # Use latest available
                    "claude-sonnet-4": "claude-3-5-sonnet-20241022",
                    "claude-3.5-sonnet": "claude-3-5-sonnet-20241022",
                    "claude-3.5-haiku": "claude-3-5-haiku-20241022"
                }
                actual_model = api_model_map.get(model, "claude-3-5-sonnet-20241022")
                
                self.llm = Anthropic(
                    model_name=actual_model,
                    anthropic_api_key=api_key,
                    temperature=0.5
                )
            elif provider == "google":
                # Map our model names to actual API model names
                api_model_map = {
                    "gemini-2.5-pro": "gemini-2.5-pro",  # Use latest available
                    "gemini-2.5-flash": "gemini-2.5-flash",
                    "gemini-2.5-flash-lite": "gemini-2.5-lite",
                }
                actual_model = api_model_map.get(model, "gemini-2.5-flash")
                
                self.llm = ChatGoogleGenerativeAI(
                    model=actual_model,
                    google_api_key=api_key,
                    temperature=0.5
                )
            else:
                raise ValueError(f"Unsupported provider: {provider}")
                
            self.current_provider = provider
            self.current_model = model
            return True
            
        except Exception as e:
            print(f"Error initializing LLM: {e}")
            self.llm = None
            return False
    
    def is_configured(self) -> bool:
        """Check if LLM is properly configured"""
        return self.llm is not None
    
    def get_current_config(self) -> Dict[str, str]:
        """Get current LLM configuration"""
        return {
            "provider": self.current_provider or "None",
            "model": self.current_model or "None",
            "status": "Configured" if self.llm else "Not Configured"
        }
    
    def extract_text_from_file(self, file_path: str, file_type: str) -> str:
        """Extract text content from various file types"""
        try:
            if file_type.lower() == 'pdf':
                return self._extract_from_pdf(file_path)
            elif file_type.lower() in ['xlsx', 'xls']:
                return self._extract_from_excel(file_path)
            elif file_type.lower() == 'csv':
                return self._extract_from_csv(file_path)
            else:
                raise ValueError(f"Unsupported file type: {file_type}")
        except Exception as e:
            raise ValueError(f"Error reading {file_type} file: {str(e)}")
    
    def _extract_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF file"""
        text = ""
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        return text
    
    def _extract_from_excel(self, file_path: str) -> str:
        """Extract text from Excel file"""
        try:
            df = pd.read_excel(file_path, sheet_name=None)
            text = ""
            for sheet_name, sheet_df in df.items():
                text += f"Sheet: {sheet_name}\n"
                text += sheet_df.to_string(index=False) + "\n\n"
            return text
        except Exception as e:
            # Fallback to openpyxl if pandas fails
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    text += "\t".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
                text += "\n"
            return text
    
    def _extract_from_csv(self, file_path: str) -> str:
        """Extract text from CSV file"""
        df = pd.read_csv(file_path)
        return df.to_string(index=False)
    
    async def analyze_financial_data(self, file_content: str, progress_callback=None) -> Optional[FinancialData]:
        """Analyze financial data using LLM"""
        if not self.llm:
            raise ValueError("LLM not configured")
        
        if progress_callback:
            progress_callback("Preparing analysis prompt...")
        
        # Setup parser and prompt
        parser = JsonOutputParser(pydantic_object=FinancialData)
        
        prompt = PromptTemplate(
            template="""You are a financial analyst expert specializing in extracting financial data for Beneish M-Score calculation.

CRITICAL INSTRUCTIONS:
1. Extract data for exactly TWO consecutive years (Year 1 = previous/older year, Year 2 = current/newer year)
2. Return financial values in millions (if stated as thousands, convert to millions by dividing by 1000)
3. Use 0 for any missing values
4. Ensure all numbers are positive (take absolute values if negative where it doesn't make sense)

REQUIRED FIELDS for each year:
- revenue (net sales/total revenue)
- cost_of_goods_sold (COGS)
- selling_general_admin_expense (SG&A expenses)
- depreciation (depreciation expense)
- net_income_continuing_operations (net income from continuing operations)
- accounts_receivables (accounts receivable/trade receivables)
- current_assets (total current assets)
- property_plant_equipment (PP&E/fixed assets)
- securities (long-term investments/marketable securities)
- total_assets (total assets)
- current_liabilities (total current liabilities)
- total_long_term_debt (long-term debt)
- cash_flow_operations (cash flow from operating activities)

Financial Document Content:
{text}

{format_instructions}

Extract the company name and financial data in the specified JSON format. Be precise with numbers and ensure consistency between years.
""",
            input_variables=["text"],
            partial_variables={"format_instructions": parser.get_format_instructions()}
        )
        
        if progress_callback:
            progress_callback("Sending data to AI for analysis...")
        
        chain = prompt | self.llm | parser
        
        try:
            result = await chain.ainvoke({"text": file_content})
            
            if progress_callback:
                progress_callback("Processing AI response...")
            
            return FinancialData(**result)
        except Exception as e:
            print(f"Error in LLM analysis: {e}")
            if progress_callback:
                progress_callback(f"Analysis failed: {str(e)}")
            return None